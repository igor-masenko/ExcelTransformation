import numpy as np
import openpyxl as xlsx
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

COLNAMES_XLSX = [
    'Код товара', 'Фирма', 'Контрагент', 'ИНН', 'Товар/Документ',
    'Начальный остаток', 'Приход', 'Расход', 'Конечный остаток',
]


def read_xlsx(path: str, sheet_name: str = None) -> Worksheet:
    wb = xlsx.load_workbook(path, data_only=True)
    if sheet_name is not None:
        return wb[sheet_name]
    else:
        return wb.active


def find_header_row(worksheet: Worksheet, headers: list) -> int:
    header_len = len(headers)
    for row_num, cells in enumerate(worksheet.rows):
        header_cells = cells[:header_len]
        if all(map(lambda x, y: x.value == y, header_cells, headers)):
            return row_num + 1
    return -1


def match_colname_to_cell(
        worksheet: Worksheet,
        headers: list,
        header_row: int,
) -> dict:
    return {k: cell for k, cell in zip(headers, worksheet[header_row])}


def excel_to_dataframe(
        workheet: Worksheet,
        colname_to_cell: dict,
        min_row: int = 0,
        max_row: int = None,
) -> pd.DataFrame:

    range_param = {
        'min_row': min_row,
        'max_row': max_row if max_row is not None else workheet.max_row,
        'min_col': min(map(lambda x: x.col_idx, colname_to_cell.values())),
        'max_col': max(map(lambda x: x.col_idx, colname_to_cell.values())),
    }

    warehouse = None
    product_name = None
    product_code = None
    open_balance = None

    cols_df = ['warehouse',
               'productCode', 'productName', 'openBalance', 'finalBalance',
               'doc', 'company', 'contractor', 'tin', 'income', 'outcome']
    d = {k: [] for k in cols_df}

    for row_ in workheet.iter_rows(**range_param):

        wp = row_[colname_to_cell['Товар/Документ'].col_idx - 1]

        if wp.font.size == 8:
            doc = wp.value

            d['warehouse'].append(warehouse)
            d['productCode'].append(product_code)
            d['productName'].append(product_name)
            d['doc'].append(doc)
            d['openBalance'].append(open_balance)

            d['company'].append(row_[colname_to_cell['Фирма'].col_idx - 1].value)
            d['contractor'].append(row_[colname_to_cell['Контрагент'].col_idx - 1].value)
            d['tin'].append(row_[colname_to_cell['ИНН'].col_idx - 1].value)
            d['income'].append(row_[colname_to_cell['Приход'].col_idx - 1].value)
            d['outcome'].append(row_[colname_to_cell['Расход'].col_idx - 1].value)
            d['finalBalance'].append(row_[colname_to_cell['Конечный остаток'].col_idx - 1].value)

        elif wp.font.size == 11:
            product_name = wp.value
            product_code = row_[colname_to_cell['Код товара'].col_idx - 1].value
            open_balance = row_[colname_to_cell['Начальный остаток'].col_idx - 1].value
        elif wp.font.size == 12:
            warehouse = wp.value
        else:
            raise ValueError(f'Неопознанный формат. Ожидался 12, 11, 8. Получено: {wp.font.size}.')

    df = pd.DataFrame(d)

    str_cols = ['warehouse', 'productName', 'doc', 'company', 'contractor', 'tin']
    df[str_cols] = df[str_cols].astype(str)

    int32 = ['productCode']
    df[int32] = df[int32].astype(np.int32)

    float32 = ['openBalance', 'finalBalance', 'income', 'outcome']
    df[float32] = df[float32].astype(np.float32)
    df[float32] = df[float32].fillna(0)

    pattern = r'(\d{,2}\.\d{,2}\.\d{4} \d{,2}:\d{,2}:\d{,2})'
    df['datetimeofdoc'] = df['doc'].str.extract(pattern)
    df['datetimeofdoc'] = pd.to_datetime(df['datetimeofdoc'])

    return df
